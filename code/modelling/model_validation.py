"""Siamese DNN — standalone holdout + validation evaluation.

Script version of model_validation.ipynb. Loads a trained `dnn_siamese_cor.pt`
checkpoint and scores it on the two out-of-training pair tables:

  - Holdout    (`holdout_*` parquet files): Bainbridge mega-study, item-
    disjoint from training but same scale families. Tests generalization to
    new items.
  - Validation (`validation_*` parquet files): SurveyBot study, entirely
    different scales (AAID + BFI-10) never seen in training. Tests
    generalization to new scale formats.

Also computes the noise ceiling (sampling precision of the split-half
targets, via data_integration.R's `split_half_correlations()` output) and the
polarity diagnostic (mean predicted r | true r < 0, sign concordance), and
appends one row to performance_logbook.xlsx.

All config defaults to whatever model_validation.ipynb currently has
hardcoded. Paths resolve relative to this script's own location, so it can be
called from anywhere:

    python model_validation.py --emb-model Qwen-Qwen3-Embedding-8B
    python model_validation.py --emb-model intfloat-e5-mistral-7b-instruct --no-plots

The notebook (model_validation.ipynb) is untouched and still works
standalone; this script is a drop-in equivalent for batch/HPC use.
"""

import argparse
import os
from datetime import datetime
from pathlib import Path

import matplotlib

matplotlib.use("Agg")  # headless: always save figures, never try to pop a window
import matplotlib.pyplot as plt
import numpy as np
import polars as pl
import seaborn as sns
import torch
from scipy import stats as sstats
from scipy.stats import pearsonr
from sklearn.metrics import mean_absolute_error, mean_squared_error, r2_score

SCRIPT_DIR = Path(__file__).resolve().parent          # code/modelling
REPO_ROOT = SCRIPT_DIR.parent.parent                   # project root


# --------------------------------------------------------------------------- #
# CLI
# --------------------------------------------------------------------------- #

def build_argparser() -> argparse.ArgumentParser:
    p = argparse.ArgumentParser(
        description="Evaluate a trained PAIR Siamese DNN checkpoint on holdout + validation.",
        formatter_class=argparse.ArgumentDefaultsHelpFormatter,
    )

    p.add_argument("--emb-model", default="Qwen-Qwen3-Embedding-8B",
                    help="Embedding backbone name (matches the folder under data/raw/ and models/).")
    p.add_argument("--data-root", default=str(REPO_ROOT / "data" / "raw"),
                    help="Root of the raw parquet data.")
    p.add_argument("--models-root", default=str(REPO_ROOT / "models"),
                    help="Root under which the checkpoint + scaler live, in a per-backbone subfolder.")
    p.add_argument("--logbook-path", default=str(REPO_ROOT / "performance_logbook.xlsx"),
                    help="xlsx file that one row per evaluation run gets appended to.")

    p.add_argument("--seed", type=int, default=42)
    p.add_argument("--encoder-dims", default="384",
                    help="Comma-separated encoder layer widths. Must match the checkpoint being loaded.")
    p.add_argument("--head-dims", default="256,130",
                    help="Comma-separated head layer widths. Must match the checkpoint being loaded.")
    p.add_argument("--dropout", type=float, default=0.163,
                    help="Inert at eval time (model.eval() disables dropout) -- kept only so the "
                         "constructed architecture matches training's module shapes exactly.")
    p.add_argument("--use-skip", action="store_true", default=False,
                    help="Must match how the checkpoint was trained.")
    p.add_argument("--eval-chunk", type=int, default=8192)

    p.add_argument("--no-plots", action="store_true", default=True,
                    help="Skip generating/saving the holdout + validation scatter plots.")
    p.add_argument("--no-log", action="store_true", default=False,
                    help="Skip appending a row to performance_logbook.xlsx.")
    p.add_argument("--notes", default="",
                    help="Free-text note stored in the logbook row (e.g. 'final model').")

    return p


def parse_dims(s: str) -> tuple:
    s = s.strip()
    if not s:
        return ()
    return tuple(int(x) for x in s.split(","))


def add_global_sim(df, emb, name_to_idx, chunk=50_000):
    known = pl.Series("item", list(name_to_idx.keys())).implode()
    df = df.filter(pl.col("Parameter1").is_in(known) & pl.col("Parameter2").is_in(known))
    i1 = np.fromiter((name_to_idx[p] for p in df["Parameter1"].to_list()), np.int64)
    i2 = np.fromiter((name_to_idx[p] for p in df["Parameter2"].to_list()), np.int64)

    En = torch.nn.functional.normalize(emb, dim=1)
    out = np.empty(len(i1), dtype=np.float32)
    with torch.no_grad():
        for s in range(0, len(i1), chunk):
            sl = slice(s, s + chunk)
            a = En[torch.as_tensor(i1[sl], device=En.device)]
            b = En[torch.as_tensor(i2[sl], device=En.device)]
            out[sl] = (a * b).sum(dim=1).cpu().numpy()
    return df.with_columns(pl.Series("global_sim", out))


def noise_ceiling(splithalf_path, label):
    sh = pl.read_parquet(splithalf_path).drop_nulls(subset=["r_half1", "r_half2"])
    r_half, _ = pearsonr(sh["r_half1"].to_numpy(), sh["r_half2"].to_numpy())
    target_consistency = (2 * r_half) / (1 + r_half)
    print(f"=== {label}: noise ceiling ===")
    print(f"N pairs (split-half table): {sh.height:,}")
    print(f"Split-half correlation (each half ~N/2): {r_half:.4f}")
    print(f"Sampling-based target consistency (full N): {target_consistency:.4f}")
    print(f"Noise ceiling, sqrt(target consistency): {np.sqrt(target_consistency):.4f}")
    return target_consistency, sh.height


def polarity_diagnostic(y_true, y_pred, label):
    neg = y_true < 0
    pos = y_true > 0

    mean_pred_given_neg = float(y_pred[neg].mean()) if neg.sum() > 0 else float("nan")
    mean_pred_given_pos = float(y_pred[pos].mean()) if pos.sum() > 0 else float("nan")

    sign_concordance_neg = float((y_pred[neg] < 0).mean()) if neg.sum() > 0 else float("nan")
    sign_concordance_pos = float((y_pred[pos] > 0).mean()) if pos.sum() > 0 else float("nan")
    sign_concordance_all = float(((y_pred > 0) == (y_true > 0)).mean())

    print(f"=== {label}: polarity diagnostic ===")
    print(f"N negative-true pairs: {neg.sum():,}  |  N positive-true pairs: {pos.sum():,}")
    print(f"Mean predicted r | true r < 0:  {mean_pred_given_neg:.4f}  "
          f"(mean true r in this subset: {y_true[neg].mean():.4f})")
    print(f"Mean predicted r | true r > 0:  {mean_pred_given_pos:.4f}  "
          f"(mean true r in this subset: {y_true[pos].mean():.4f})")
    print(f"Sign concordance, true r < 0 (predicted also negative): {sign_concordance_neg:.1%}")
    print(f"Sign concordance, true r > 0 (predicted also positive): {sign_concordance_pos:.1%}")
    print(f"Sign concordance, overall: {sign_concordance_all:.1%}")

    return {
        "mean_pred_given_neg": mean_pred_given_neg,
        "mean_pred_given_pos": mean_pred_given_pos,
        "sign_concordance_neg": sign_concordance_neg,
        "sign_concordance_pos": sign_concordance_pos,
        "sign_concordance_all": sign_concordance_all,
    }


def main(argv=None):
    args = build_argparser().parse_args(argv)

    encoder_dims = parse_dims(args.encoder_dims)
    head_dims = parse_dims(args.head_dims)

    data_root = Path(args.data_root)
    models_root = Path(args.models_root)
    model_dir = models_root / args.emb_model

    train_pair_path = data_root / "item_correlations.parquet"
    train_emb_path = data_root / args.emb_model / "embeddings_raw.parquet"
    hold_pair_path = data_root / "holdout_item_correlations.parquet"
    hold_emb_path = data_root / args.emb_model / "holdout_embeddings_raw.parquet"
    val_pair_path = data_root / "validation_item_correlations.parquet"
    val_emb_path = data_root / args.emb_model / "validation_embeddings_raw.parquet"
    ckpt_path = model_dir / "dnn_siamese_cor.pt"
    scaler_path = model_dir / "quantile_transformer.joblib"

    device = torch.device("cuda:0" if torch.cuda.is_available() else "cpu")
    print(f"Device: {device}")
    torch.manual_seed(args.seed)
    np.random.seed(args.seed)

    # ---- 2. Load training embeddings + pair table (for the item lookup) ----
    emb_df = pl.read_parquet(train_emb_path)
    emb_cols = [c for c in emb_df.columns if c.startswith("emb")]
    emb_dim = len(emb_cols)
    print(f"Training embeddings: {emb_df.shape}  emb_dim={emb_dim}")

    item_to_idx = {name: i for i, name in enumerate(emb_df["item"].to_list())}
    item_emb = torch.tensor(emb_df.select(emb_cols).to_numpy(), dtype=torch.float32).to(device)
    print(f"ITEM_EMB on device: {tuple(item_emb.shape)}")

    dat = pl.read_parquet(train_pair_path).filter(
        (pl.col("r").is_not_null()) & (pl.col("r") != 1)
    )
    dat = add_global_sim(dat, item_emb, item_to_idx)
    print(f"Pair rows (with embeddings): {dat.height:,}")

    # ---- 3. Load the train-fitted preprocessor ----
    import joblib

    scaler = joblib.load(scaler_path)
    aux_dim = scaler.n_features_in_
    print(f"Loaded scaler from {scaler_path}")
    print(f"AUX_DIM (numeric) = {aux_dim}")

    # ---- 4. Model architecture + checkpoint ----
    from siamese_model import SiameseDNN

    model = SiameseDNN(
        emb_dim=emb_dim, aux_dim=aux_dim,
        encoder_dims=encoder_dims, head_dims=head_dims,
        dropout=args.dropout, use_skip=args.use_skip,
    ).to(device)
    model.load_state_dict(torch.load(ckpt_path, map_location=device))
    model.eval()
    print(f"Loaded checkpoint from {ckpt_path}")
    print(f"Trainable params: {sum(p.numel() for p in model.parameters() if p.requires_grad):,}")

    # ---- 5. Holdout evaluation ----
    hold_df = pl.read_parquet(hold_pair_path).filter(
        (pl.col("r").is_not_null()) & (pl.col("r") != 1)
    )
    print(f"\nHoldout pair rows on disk: {hold_df.height:,}")

    hold_emb_df = pl.read_parquet(hold_emb_path)
    hold_emb_cols = [c for c in hold_emb_df.columns if c.startswith("emb")]
    assert len(hold_emb_cols) == emb_dim, "Holdout embedding dim doesn't match training dim"

    combined_item_to_idx = dict(item_to_idx)
    new_local_rows, new_names = [], []
    for local_row, name in enumerate(hold_emb_df["item"].to_list()):
        if name not in combined_item_to_idx:
            combined_item_to_idx[name] = len(item_to_idx) + len(new_local_rows)
            new_local_rows.append(local_row)
            new_names.append(name)

    if new_local_rows:
        hold_emb_np = hold_emb_df.select(hold_emb_cols).to_numpy()[new_local_rows]
        hold_emb_new = torch.tensor(hold_emb_np, dtype=torch.float32).to(device)
        combined_emb = torch.cat([item_emb, hold_emb_new], dim=0)
    else:
        combined_emb = item_emb
    print(f"Combined embedding lookup: {len(item_to_idx):,} train + {len(new_names):,} new holdout = "
          f"{combined_emb.shape[0]:,} items")

    known_all = pl.Series("item", list(combined_item_to_idx.keys())).implode()
    pre_n = hold_df.height
    hold_df = hold_df.filter(
        pl.col("Parameter1").is_in(known_all) & pl.col("Parameter2").is_in(known_all)
    )
    print(f"After item-lookup filter: {hold_df.height:,} pairs (dropped {pre_n - hold_df.height:,})")

    hold_df = add_global_sim(hold_df, combined_emb, combined_item_to_idx)
    idx1 = np.fromiter((combined_item_to_idx[p] for p in hold_df["Parameter1"].to_list()), np.int64)
    idx2 = np.fromiter((combined_item_to_idx[p] for p in hold_df["Parameter2"].to_list()), np.int64)

    aux = scaler.transform(hold_df.select(pl.col("global_sim")).to_numpy())
    y_h = hold_df.select("r").to_numpy().flatten().astype(np.float32)

    idx1_t = torch.tensor(idx1, dtype=torch.long, device=device)
    idx2_t = torch.tensor(idx2, dtype=torch.long, device=device)
    aux_t = torch.tensor(aux, dtype=torch.float32, device=device)

    chunks = []
    with torch.no_grad():
        for s in range(0, idx1_t.shape[0], args.eval_chunk):
            e1 = combined_emb[idx1_t[s:s + args.eval_chunk]]
            e2 = combined_emb[idx2_t[s:s + args.eval_chunk]]
            ax_ = aux_t[s:s + args.eval_chunk]
            chunks.append(torch.tanh(model(e1, e2, ax_)).cpu().numpy())
    h_preds = np.concatenate(chunks)

    corr_h, _ = pearsonr(y_h, h_preds)
    rmse_h = np.sqrt(mean_squared_error(y_h, h_preds))
    r2_h = r2_score(y_h, h_preds)
    mae_h = mean_absolute_error(y_h, h_preds)

    print("\n=== SIAMESE DNN HOLDOUT EVALUATION ===")
    print(f"N pairs:    {len(y_h):,}")
    print(f"Pearson r:  {corr_h:.4f}")
    print(f"R-squared:  {r2_h:.4f}")
    print(f"RMSE:       {rmse_h:.4f}")
    print(f"MAE:        {mae_h:.4f}")

    if not args.no_plots:
        fig, ax = plt.subplots(figsize=(8, 8))
        sns.regplot(x=y_h, y=h_preds, ax=ax, scatter_kws={"alpha": 0.3, "color": "black"},
                    line_kws={"color": "red"})
        ax.plot([-1, 1], [-1, 1], color="blue", linestyle="--")
        ax.set_xlim(-1, 1)
        ax.set_ylim(-1, 1)
        ax.set_aspect("equal", adjustable="box")
        ax.set_title(f"Holdout: Actual vs Predicted (r = {corr_h:.3f})  n = {len(y_h):,}")
        ax.set_xlabel("Human Correlation (Pearson r)")
        ax.set_ylabel("Model Prediction")
        ax.grid(True, alpha=0.3)
        plt.tight_layout()
        plot_path = model_dir / "holdout_scatter_eval.png"
        plt.savefig(plot_path, dpi=150)
        plt.close(fig)
        print(f"Saved holdout scatter -> {plot_path}")

    # ---- 6. Validation evaluation ----
    val_df = pl.read_parquet(val_pair_path).filter(
        (pl.col("r").is_not_null()) & (pl.col("r") != 1)
    )
    print(f"\nValidation pair rows on disk: {val_df.height:,}")

    val_emb_df = pl.read_parquet(val_emb_path)
    val_emb_cols = [c for c in val_emb_df.columns if c.startswith("emb")]
    assert len(val_emb_cols) == emb_dim, "Validation embedding dim doesn't match training dim"

    val_item_to_idx = dict(combined_item_to_idx)
    new_local_rows_v, new_names_v = [], []
    for local_row, name in enumerate(val_emb_df["item"].to_list()):
        if name not in val_item_to_idx:
            val_item_to_idx[name] = combined_emb.shape[0] + len(new_local_rows_v)
            new_local_rows_v.append(local_row)
            new_names_v.append(name)

    if new_local_rows_v:
        val_emb_np = val_emb_df.select(val_emb_cols).to_numpy()[new_local_rows_v]
        val_emb_new = torch.tensor(val_emb_np, dtype=torch.float32).to(device)
        val_combined_emb = torch.cat([combined_emb, val_emb_new], dim=0)
    else:
        val_combined_emb = combined_emb
    print(f"Validation embedding lookup: {combined_emb.shape[0]:,} prior + "
          f"{len(new_names_v):,} new validation = {val_combined_emb.shape[0]:,} items")

    known_all_v = pl.Series("item", list(val_item_to_idx.keys())).implode()
    pre_n = val_df.height
    val_df = val_df.filter(
        pl.col("Parameter1").is_in(known_all_v) & pl.col("Parameter2").is_in(known_all_v)
    )
    print(f"After item-lookup filter: {val_df.height:,} pairs (dropped {pre_n - val_df.height:,})")

    val_df = add_global_sim(val_df, val_combined_emb, val_item_to_idx)
    idx1_v = np.fromiter((val_item_to_idx[p] for p in val_df["Parameter1"].to_list()), np.int64)
    idx2_v = np.fromiter((val_item_to_idx[p] for p in val_df["Parameter2"].to_list()), np.int64)

    aux_v = scaler.transform(val_df.select(pl.col("global_sim")).to_numpy())
    y_v = val_df.select("r").to_numpy().flatten().astype(np.float32)

    idx1_vt = torch.tensor(idx1_v, dtype=torch.long, device=device)
    idx2_vt = torch.tensor(idx2_v, dtype=torch.long, device=device)
    aux_vt = torch.tensor(aux_v, dtype=torch.float32, device=device)

    chunks_v = []
    model.eval()
    with torch.no_grad():
        for s in range(0, idx1_vt.shape[0], args.eval_chunk):
            e1 = val_combined_emb[idx1_vt[s:s + args.eval_chunk]]
            e2 = val_combined_emb[idx2_vt[s:s + args.eval_chunk]]
            ax_ = aux_vt[s:s + args.eval_chunk]
            chunks_v.append(torch.tanh(model(e1, e2, ax_)).cpu().numpy())
    v_preds = np.concatenate(chunks_v)

    corr_v, _ = pearsonr(y_v, v_preds)
    rmse_v = np.sqrt(mean_squared_error(y_v, v_preds))
    r2_v = r2_score(y_v, v_preds)
    mae_v = mean_absolute_error(y_v, v_preds)

    print("\n=== SIAMESE DNN VALIDATION EVALUATION ===")
    print(f"N pairs:    {len(y_v):,}")
    print(f"Pearson r:  {corr_v:.4f}")
    print(f"R-squared:  {r2_v:.4f}")
    print(f"RMSE:       {rmse_v:.4f}")
    print(f"MAE:        {mae_v:.4f}")

    if not args.no_plots:
        x = np.asarray(y_v, dtype=float)
        y = np.asarray(v_preds, dtype=float)
        n = len(x)
        order = np.argsort(x)
        xs, ys = x[order], y[order]
        W = max(200, int(0.05 * n))
        step = max(1, W // 5)

        gx, glo, gmed, ghi = [], [], [], []
        for c in range(W // 2, n - W // 2, step):
            sl = slice(c - W // 2, c + W // 2)
            gx.append(xs[c])
            glo.append(np.quantile(ys[sl], 0.025))
            gmed.append(np.median(ys[sl]))
            ghi.append(np.quantile(ys[sl], 0.975))
        gx, glo, gmed, ghi = map(np.array, (gx, glo, gmed, ghi))

        b1, b0 = np.polyfit(x, y, 1)
        xg = np.linspace(-1, 1, 200)

        fig, ax = plt.subplots(figsize=(8, 8))
        ax.scatter(x, y, alpha=0.3, color="black", s=10)
        ax.fill_between(gx, glo, ghi, color="red", alpha=0.2, label="empirical 95% band (x-local)")
        ax.plot(gx, gmed, color="red", lw=2, label="local median")
        ax.plot(xg, b0 + b1 * xg, color="orange", lw=1.5, ls=":", label=f"OLS fit (slope {b1:.3f})")
        ax.plot([-1, 1], [-1, 1], color="blue", linestyle="--", label="y = x")
        ax.set_xlim(-1, 1)
        ax.set_ylim(-1, 1)
        ax.set_aspect("equal", adjustable="box")
        ax.set_title(f"Validation: Actual vs Predicted (r = {corr_v:.3f})  n = {n:,}")
        ax.set_xlabel("Human Correlation (Pearson r)")
        ax.set_ylabel("Model Prediction")
        ax.legend(loc="upper left", fontsize=8)
        ax.grid(True, alpha=0.3)
        plt.tight_layout()
        plot_path = model_dir / "validation_scatter_eval.png"
        plt.savefig(plot_path, dpi=150)
        plt.close(fig)
        print(f"Saved validation scatter -> {plot_path}")

    # ---- 7. Noise ceiling ----
    print()
    holdout_target_consistency, n_splithalf_holdout = noise_ceiling(
        data_root / "holdout_item_correlations_splithalf.parquet", "Holdout (Bainbridge)"
    )
    holdout_noise_ceiling = np.sqrt(holdout_target_consistency)
    holdout_r_noise_aware = corr_h / holdout_noise_ceiling
    print(f"\nPAIR holdout r = {corr_h:.4f}  |  noise ceiling = {holdout_noise_ceiling:.4f}  |  "
          f"noise-aware r = {holdout_r_noise_aware:.4f}")

    print()
    validation_target_consistency, n_splithalf_val = noise_ceiling(
        data_root / "validation_item_correlations_splithalf.parquet", "Validation (SurveyBot3000)"
    )
    validation_noise_ceiling = np.sqrt(validation_target_consistency)
    validation_r_noise_aware = corr_v / validation_noise_ceiling
    print(f"\nPAIR validation r = {corr_v:.4f}  |  noise ceiling = {validation_noise_ceiling:.4f}  |  "
          f"noise-aware r = {validation_r_noise_aware:.4f}")

    # ---- 8. Polarity diagnostic ----
    print()
    holdout_polarity = polarity_diagnostic(y_h, h_preds, "Holdout (Bainbridge)")
    print()
    val_polarity = polarity_diagnostic(y_v, v_preds, "Validation (SurveyBot3000)")

    # ---- 9. Log to performance logbook ----
    if not args.no_log:
        import openpyxl  # noqa: F401  (import kept for a clear error message if missing)
        from openpyxl import Workbook, load_workbook

        columns = [
            "timestamp", "embedding_model", "checkpoint_path",
            "encoder_dims", "head_dims", "dropout", "emb_dim", "n_train_items",
            "n_holdout", "holdout_r", "holdout_r2", "holdout_rmse", "holdout_mae",
            "holdout_slope", "holdout_intercept",
            "holdout_target_consistency", "holdout_noise_ceiling", "holdout_r_noise_aware",
            "holdout_mean_pred_given_neg", "holdout_sign_concordance",
            "n_validation", "val_r", "val_r2", "val_rmse", "val_mae",
            "val_slope", "val_intercept",
            "val_target_consistency", "val_noise_ceiling", "val_r_noise_aware",
            "val_mean_pred_given_neg", "val_sign_concordance",
            "notes",
        ]

        h_slope, h_intercept, *_ = sstats.linregress(y_h, h_preds)
        v_slope, v_intercept, *_ = sstats.linregress(y_v, v_preds)

        row = [
            datetime.now().isoformat(timespec="seconds"),
            args.emb_model,
            str(ckpt_path),
            str(encoder_dims),
            str(head_dims),
            args.dropout,
            emb_dim,
            len(item_to_idx),
            len(y_h), corr_h, r2_h, rmse_h, mae_h, h_slope, h_intercept,
            holdout_target_consistency, holdout_noise_ceiling, holdout_r_noise_aware,
            holdout_polarity["mean_pred_given_neg"], holdout_polarity["sign_concordance_all"],
            len(y_v), corr_v, r2_v, rmse_v, mae_v, v_slope, v_intercept,
            validation_target_consistency, validation_noise_ceiling, validation_r_noise_aware,
            val_polarity["mean_pred_given_neg"], val_polarity["sign_concordance_all"],
            args.notes,
        ]

        logbook_path = Path(args.logbook_path)
        if logbook_path.exists():
            wb = load_workbook(logbook_path)
            ws = wb.active
        else:
            wb = Workbook()
            ws = wb.active
            ws.title = "performance_log"
            ws.append(columns)

        ws.append(row)
        wb.save(logbook_path)

        print(f"\nLogged evaluation to {logbook_path}")
        print(f"  {args.emb_model} | holdout r={corr_h:.4f} | validation r={corr_v:.4f}")

    return {
        "holdout": {"n": len(y_h), "r": corr_h, "r2": r2_h, "rmse": rmse_h, "mae": mae_h,
                    "noise_ceiling": holdout_noise_ceiling, "r_noise_aware": holdout_r_noise_aware},
        "validation": {"n": len(y_v), "r": corr_v, "r2": r2_v, "rmse": rmse_v, "mae": mae_v,
                       "noise_ceiling": validation_noise_ceiling, "r_noise_aware": validation_r_noise_aware},
    }


if __name__ == "__main__":
    main()
